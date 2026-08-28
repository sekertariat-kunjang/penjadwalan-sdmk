import io
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, Pegawai, Ruangan, Shift, Jadwal, StatusJadwal

INDONESIAN_DAYS = ['Sen', 'Sel', 'Rab', 'Kam', 'Jum', 'Sab', 'Min']
INDONESIAN_MONTHS = [
    '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
]

def generate_excel_schedule(year, month):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal SDMK {INDONESIAN_MONTHS[month]} {year}"
    ws.views.sheetView[0].showGridLines = True

    # Check Approval Status
    st_obj = StatusJadwal.query.filter_by(tahun=year, bulan=month).first()
    status_str = st_obj.status if st_obj else 'DRAFT'
    approved_by = st_obj.approved_by if st_obj else 'Kepala Puskesmas'
    approved_at = st_obj.approved_at if st_obj else ''

    # Colors & Fills
    HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    SUBHEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True, color="E2E8F0")

    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="0F172A")
    SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="475569")
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Write Header Title
    ws.merge_cells("A1:AG1")
    ws["A1"] = "PUSKESMAS SDMK - JADWAL PELAYANAN & PIKET SDMK"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    status_badge_text = f"[ STATUS: {status_str} ]"
    if status_str == 'FINAL':
        status_badge_text = f"[ STATUS RESMI: FINAL APPROVED BY KEPALA PUSKESMAS ]"

    ws.merge_cells("A2:AG2")
    ws["A2"] = f"Periode: {INDONESIAN_MONTHS[month]} {year} | {status_badge_text}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    num_days = calendar.monthrange(year, month)[1]

    # 2. Table Headers (Row 4 & Row 5)
    ws.cell(row=4, column=1, value="No").fill = HEADER_FILL
    ws.cell(row=4, column=1).font = HEADER_FONT
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A4:A5")

    ws.cell(row=4, column=2, value="Ruangan / Layanan (Klaster)").fill = HEADER_FILL
    ws.cell(row=4, column=2).font = HEADER_FONT
    ws.cell(row=4, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B4:B5")

    ws.cell(row=4, column=3, value="Pegawai Bertugas").fill = HEADER_FILL
    ws.cell(row=4, column=3).font = HEADER_FONT
    ws.cell(row=4, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C4:C5")

    # Dates columns
    for day in range(1, num_days + 1):
        col_idx = 3 + day
        dt = datetime(year, month, day)
        day_name = INDONESIAN_DAYS[dt.weekday()]
        
        cell_day = ws.cell(row=4, column=col_idx, value=day_name)
        cell_day.fill = HEADER_FILL
        cell_day.font = HEADER_FONT
        cell_day.alignment = Alignment(horizontal="center", vertical="center")

        cell_num = ws.cell(row=5, column=col_idx, value=day)
        if dt.weekday() == 6:
            cell_num.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
            cell_num.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        else:
            cell_num.fill = SUBHEADER_FILL
            cell_num.font = SUBHEADER_FONT
        cell_num.alignment = Alignment(horizontal="center", vertical="center")

    # Summary Columns Header
    start_summary_col = 3 + num_days + 1
    summary_headers = ['Total Pagi', 'Total Siang', 'Total Malam', 'Total Libur/Cuti']
    for idx, sh in enumerate(summary_headers):
        c_idx = start_summary_col + idx
        cell_s = ws.cell(row=4, column=c_idx, value=sh)
        cell_s.fill = HEADER_FILL
        cell_s.font = HEADER_FONT
        cell_s.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=4, start_column=c_idx, end_row=5, end_column=c_idx)

    # 3. Load Data & Write Matrix
    ruangan_list = Ruangan.query.order_by(Ruangan.urutan.asc()).all()

    start_str = f"{year}-{month:02d}-01"
    end_str = f"{year}-{month:02d}-{num_days:02d}"
    jadwals = Jadwal.query.filter(Jadwal.tanggal >= start_str, Jadwal.tanggal <= end_str).all()
    
    schedule_map = {}
    for j in jadwals:
        k = (j.ruangan_id, j.tanggal)
        if k not in schedule_map:
            schedule_map[k] = []
        schedule_map[k].append(j)

    current_row = 6
    no = 1
    current_klaster = None

    for r in ruangan_list:
        if r.klaster != current_klaster:
            current_klaster = r.klaster
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=start_summary_col + len(summary_headers) - 1)
            group_cell = ws.cell(row=current_row, column=1, value=f"--- {current_klaster.upper()} ---")
            group_cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
            group_cell.font = Font(name="Calibri", size=10, bold=True, color="F8FAFC")
            group_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        ws.cell(row=current_row, column=1, value=no).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=1).border = THIN_BORDER
        no += 1

        ws.cell(row=current_row, column=2, value=r.nama).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=current_row, column=2).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=current_row, column=2).border = THIN_BORDER

        room_jadwals = [j for j in jadwals if j.ruangan_id == r.id]
        staff_names = sorted(list(set(j.pegawai.nama for j in room_jadwals if j.pegawai)))
        staff_name_str = ", ".join(staff_names[:3]) + (f" (+{len(staff_names)-3})" if len(staff_names) > 3 else "") if staff_names else "-"

        ws.cell(row=current_row, column=3, value=staff_name_str).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=current_row, column=3).font = Font(name="Calibri", size=9)
        ws.cell(row=current_row, column=3).border = THIN_BORDER

        count_pagi = 0
        count_siang = 0
        count_malam = 0
        count_libur = 0

        for day in range(1, num_days + 1):
            col_idx = 3 + day
            tgl_str = f"{year}-{month:02d}-{day:02d}"
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            list_j = schedule_map.get((r.id, tgl_str), [])
            if list_j:
                codes = [j.shift.kode for j in list_j if j.shift]
                cell.value = ", ".join(codes)
                
                first_j = list_j[0]
                if first_j.shift:
                    hex_color = first_j.shift.warna_bg.replace('#', '')
                    if len(hex_color) == 6:
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                
                for j in list_j:
                    if j.shift:
                        sc = j.shift.kode
                        if sc == 'P': count_pagi += 1
                        elif sc == 'S': count_siang += 1
                        elif sc == 'M': count_malam += 1
                        elif sc in ['L', 'C']: count_libur += 1
            else:
                cell.value = "-"
                cell.font = Font(name="Calibri", size=9, color="94A3B8")

        # Fill Summaries
        ws.cell(row=current_row, column=start_summary_col, value=count_pagi).border = THIN_BORDER
        ws.cell(row=current_row, column=start_summary_col).alignment = Alignment(horizontal="center")
        
        ws.cell(row=current_row, column=start_summary_col + 1, value=count_siang).border = THIN_BORDER
        ws.cell(row=current_row, column=start_summary_col + 1).alignment = Alignment(horizontal="center")
        
        ws.cell(row=current_row, column=start_summary_col + 2, value=count_malam).border = THIN_BORDER
        ws.cell(row=current_row, column=start_summary_col + 2).alignment = Alignment(horizontal="center")
        
        ws.cell(row=current_row, column=start_summary_col + 3, value=count_libur).border = THIN_BORDER
        ws.cell(row=current_row, column=start_summary_col + 3).alignment = Alignment(horizontal="center")

        current_row += 1
        no += 1

    # 4. Signature Block
    current_row += 2
    ws.cell(row=current_row, column=2, value="Mengetahui,").font = Font(name="Calibri", size=10)
    ws.cell(row=current_row, column=start_summary_col - 5, value=f"Puskesmas, {num_days} {INDONESIAN_MONTHS[month]} {year}").font = Font(name="Calibri", size=10)
    
    current_row += 1
    ws.cell(row=current_row, column=2, value="Kepala Puskesmas").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=current_row, column=start_summary_col - 5, value="Pengelola SDMK / Kasubag TU").font = Font(name="Calibri", size=10, bold=True)

    if status_str == 'FINAL':
        current_row += 1
        ws.cell(row=current_row, column=2, value=f"[ DISETUJUI RESMI: {approved_at} ]").font = Font(name="Calibri", size=9, bold=True, color="047857")

    current_row += 3
    ws.cell(row=current_row, column=2, value=f"( {approved_by} )").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=current_row, column=start_summary_col - 5, value="( ___________________________ )").font = Font(name="Calibri", size=10, bold=True)

    # Column Width Auto Adjustment
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    for day in range(1, num_days + 1):
        col_letter = get_column_letter(3 + day)
        ws.column_dimensions[col_letter].width = 4.5
    
    for idx in range(len(summary_headers)):
        col_letter = get_column_letter(start_summary_col + idx)
        ws.column_dimensions[col_letter].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
