import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, Pegawai, Ruangan, Shift, User, StatusJadwal, Jadwal
from datetime import datetime

def generate_excel_template(filepath="template_sdmk_puskesmas.xlsx"):
    """
    Menghasilkan file template Excel (.xlsx) rapi untuk diisi oleh pengguna
    dengan data real Pegawai dan Ruangan / Layanan Puskesmas.
    """
    wb = openpyxl.Workbook()

    # Style definitions
    header_fill = PatternFill(start_color="115E59", end_color="115E59", fill_type="solid") # Dark Teal
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid") # Light Emerald
    sub_font = Font(name="Calibri", size=10, italic=True, color="065F46")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # -------------------------------------------------------------
    # SHEET 1: Data Pegawai
    # -------------------------------------------------------------
    ws_pegawai = wb.active
    ws_pegawai.title = "Data Pegawai"
    ws_pegawai.views.sheetView[0].showGridLines = True

    # Description row
    ws_pegawai.merge_cells("A1:D1")
    ws_pegawai["A1"] = "PETUNJUK: Isikan daftar pegawai SDMK Puskesmas di bawah ini. NIP dan No HP boleh dikosongkan jika belum ada."
    ws_pegawai["A1"].font = sub_font
    ws_pegawai["A1"].fill = sub_fill
    ws_pegawai["A1"].alignment = align_left

    headers_pegawai = ["NIP", "Nama Lengkap & Gelar", "Profesi", "No. HP / WhatsApp"]
    for col_num, h in enumerate(headers_pegawai, 1):
        cell = ws_pegawai.cell(row=2, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Sample rows
    sample_pegawai = [
        ["198503122010011001", "dr. Ahmad Fauzi, Sp.KKLP", "Dokter Umum", "081234567801"],
        ["198807202014022003", "dr. Siti Rahmawati", "Dokter Umum", "081234567802"],
        ["199011052016031005", "drg. Maya Indah", "Dokter Gigi", "081234567803"],
        ["198904152012012004", "Bdn. Rina Amalia, S.Tr.Keb", "Bidan", "081234567804"],
        ["198709182011011002", "Ns. Hendra Kurniawan, S.Kep", "Perawat", "081234567807"],
        ["198610112009022001", "apt. Fitriani, S.Farm", "Apoteker", "081234567811"],
        ["199102282015031007", "Rahmat Hidayat, Amd.AK", "Pranata Laboratorium", "081234567813"],
        ["198405122008011003", "Agus Wijaya", "Staf TU", "081234567816"]
    ]

    for row_idx, data in enumerate(sample_pegawai, 3):
        for col_idx, val in enumerate(data, 1):
            cell = ws_pegawai.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = align_center if col_idx in [1, 3, 4] else align_left

    # Auto column width
    for col in ws_pegawai.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_pegawai.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # -------------------------------------------------------------
    # SHEET 2: Data Ruangan & Layanan
    # -------------------------------------------------------------
    ws_ruang = wb.create_sheet(title="Data Ruangan")
    ws_ruang.views.sheetView[0].showGridLines = True

    ws_ruang.merge_cells("A1:D1")
    ws_ruang["A1"] = "PETUNJUK: Isikan unit layanan Puskesmas. Klaster disesuaikan dengan ILP (Klaster 1 s/d 4, Lintas Klaster, Luar Induk)."
    ws_ruang["A1"].font = sub_font
    ws_ruang["A1"].fill = sub_fill
    ws_ruang["A1"].alignment = align_left

    headers_ruang = ["Nama Layanan / Ruangan", "Kode Singkat", "Kategori Klaster", "Urutan Tampil"]
    for col_num, h in enumerate(headers_ruang, 1):
        cell = ws_ruang.cell(row=2, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    sample_ruang = [
        ["Pendaftaran & Rekam Medis", "RM", "Klaster 1 (Manajemen)", 1],
        ["Tata Usaha / Kasir", "TU", "Klaster 1 (Manajemen)", 2],
        ["Poli KIA & KB", "KIA", "Klaster 2 (Ibu & Anak)", 3],
        ["Poli Anak & Imunisasi", "IMN", "Klaster 2 (Ibu & Anak)", 4],
        ["Poli Umum", "UMM", "Klaster 3 (Dewasa & Lansia)", 5],
        ["Poli Gigi & Mulut", "GGI", "Klaster 3 (Dewasa & Lansia)", 6],
        ["Poli Lansia & PTM", "PTM", "Klaster 3 (Dewasa & Lansia)", 7],
        ["Farmasi / Apotek", "APT", "Klaster 4 (P2)", 8],
        ["Laboratorium", "LAB", "Klaster 4 (P2)", 9],
        ["Ruang Gizi & Kesling", "GZI", "Klaster 4 (P2)", 10],
        ["IGD / Gawat Darurat (24 Jam)", "IGD", "Klaster 5 (Lintas Klaster)", 11],
        ["Ruang Rawat Inap", "RNI", "Klaster 5 (Lintas Klaster)", 12],
        ["Pustu Desa Sukamaju", "PST1", "Luar Induk", 13],
        ["Poskesdes Harapan", "POS1", "Luar Induk", 14],
        ["Puskesmas Keliling (Pusling)", "PSL", "Luar Induk", 15]
    ]

    for row_idx, data in enumerate(sample_ruang, 3):
        for col_idx, val in enumerate(data, 1):
            cell = ws_ruang.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = align_center if col_idx in [2, 4] else align_left

    for col in ws_ruang.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ruang.column_dimensions[col_letter].width = max(max_len + 4, 18)

    wb.save(filepath)
    print(f"--> Template Excel berhasil dibuat di: {filepath}")
    return filepath


def parse_and_import_excel(filepath, reset_jadwal=True):
    """
    Membaca data Pegawai dan Ruangan dari berkas Excel (.xlsx)
    dan mengisikannya ke database SQLite.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    pegawai_imported = 0
    ruangan_imported = 0

    # 1. Parse Sheet "Data Pegawai"
    sheet_pegawai_name = next((s for s in wb.sheetnames if "pegawai" in s.lower()), wb.sheetnames[0])
    ws_peg = wb[sheet_pegawai_name]

    new_pegawai_list = []
    # Loop over rows starting from row 3 (ignoring petunjuk & header)
    for row in ws_peg.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue
        nip = str(row[0]).strip() if row[0] is not None else None
        nama = str(row[1]).strip() if row[1] is not None else None
        profesi = str(row[2]).strip() if row[2] is not None else "Staf"
        no_hp = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None

        if not nama or nama.lower().startswith("petunjuk") or nama.lower().startswith("nama"):
            continue

        new_pegawai_list.append({
            'nip': nip if nip and nip != "None" else "-",
            'nama': nama,
            'profesi': profesi if profesi and profesi != "None" else "Staf",
            'no_hp': no_hp if no_hp and no_hp != "None" else "-"
        })

    # 2. Parse Sheet "Data Ruangan"
    sheet_ruang_name = next((s for s in wb.sheetnames if "ruang" in s.lower() or "layanan" in s.lower()), None)
    new_ruangan_list = []
    if sheet_ruang_name:
        ws_rng = wb[sheet_ruang_name]
        for row in ws_rng.iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            nama_ruang = str(row[0]).strip() if row[0] is not None else None
            kode = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            klaster = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "Klaster 1 (Manajemen)"
            try:
                urutan = int(row[3]) if len(row) > 3 and row[3] is not None else 99
            except ValueError:
                urutan = 99

            if not nama_ruang or nama_ruang.lower().startswith("petunjuk") or nama_ruang.lower().startswith("nama"):
                continue

            new_ruangan_list.append({
                'nama': nama_ruang,
                'kode': kode if kode and kode != "None" else "",
                'klaster': klaster if klaster and klaster != "None" else "Klaster 1 (Manajemen)",
                'urutan': urutan
            })

    # Update Database
    if new_pegawai_list or new_ruangan_list:
        if reset_jadwal:
            Jadwal.query.delete()
            User.query.filter(User.role == 'pegawai').delete()
            if new_pegawai_list:
                Pegawai.query.delete()
            if new_ruangan_list:
                Ruangan.query.delete()

        if new_pegawai_list:
            for p_data in new_pegawai_list:
                existing = Pegawai.query.filter_by(nama=p_data['nama']).first()
                if not existing:
                    peg = Pegawai(
                        nama=p_data['nama'],
                        profesi=p_data['profesi'],
                        nip=p_data['nip'],
                        no_hp=p_data['no_hp']
                    )
                    db.session.add(peg)
                    pegawai_imported += 1
                else:
                    existing.nip = p_data['nip']
                    existing.profesi = p_data['profesi']
                    existing.no_hp = p_data['no_hp']
                    pegawai_imported += 1

        if new_ruangan_list:
            for r_data in new_ruangan_list:
                existing_r = Ruangan.query.filter_by(nama=r_data['nama']).first()
                if not existing_r:
                    rng = Ruangan(
                        nama=r_data['nama'],
                        klaster=r_data['klaster'],
                        kode=r_data['kode'],
                        urutan=r_data['urutan']
                    )
                    db.session.add(rng)
                    ruangan_imported += 1
                else:
                    existing_r.kode = r_data['kode']
                    existing_r.klaster = r_data['klaster']
                    existing_r.urutan = r_data['urutan']
                    ruangan_imported += 1

        db.session.commit()

    return {
        'status': 'success',
        'pegawai_imported': pegawai_imported,
        'ruangan_imported': ruangan_imported
    }

if __name__ == '__main__':
    generate_excel_template()
